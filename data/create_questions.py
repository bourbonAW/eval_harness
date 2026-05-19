from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

HEADERS = [
    "用户输入的问题", "期望的回复", "关联的政策", "关联的文档",
    "申报入口", "是否多意图", "知识类型", "问题是否违禁", "备注",
]

POLICY_URL = "https://gxj.gz.gov.cn/gkmlpt/content/9/9633/post_9633590.html#16649"
DOC_URL = "https://q9jvw0u5f5.feishu.cn/file/XXnvbhiOEoHVDpxm5ymcX0IFnFc"
DOC_NAME = "附件2：支持省级企业技术中心开展创新能力建设项目申报指南.docx"

def row(question, expected, intent="单意图", ktype="文档", prohibited="正常", note=""):
    return [question, expected, POLICY_URL, DOC_URL, DOC_NAME, intent, ktype, prohibited, note]

ROWS = [
    row(
        "省级企业技术中心项目专项资金支持的是哪些企业？",
        "已认定为省级企业技术中心（不含建筑业）的企业（含已获得过省级企业技术中心专题财政资金支持的企业）。"
        "项目承担单位应在广州市行政区域内设立、登记，具有独立法人资格，或为省属企业。（具体以每年申报通知为准）",
    ),
    row(
        "省级企业技术中心项目专项资金支持哪些方向项目？",
        "围绕20个战略性产业集群等重点领域，对企业在提升自主创新能力和核心竞争力过程中，依托省级企业技术中心"
        "在试验、检验检测、中试孵化、成果转化、产业应用等方面开展的项目给予支持。鼓励企业围绕技术创新的薄弱环节，"
        "加大创新研发投入，完善创新基础设施，实施创新能力建设项目。（具体以每年申报通知为准）",
    ),
    row(
        "省级企业技术中心项目专项资金支持方式及标准是？",
        "省财政资金奖补比例不超过项目新购置研发仪器设备（含配套软件）总额的40%（不含税），"
        "单个项目获得资金原则上不少于200万元、不超过800万元，同等条件下可优先支持上一年度年度评价结果为良好及以上的"
        "省级企业技术中心所在企业以及国家技术创新示范企业、国家产业技术基础公共服务平台企业、省级以上制造业创新中心"
        "牵头企业、省级以上制造业单项冠军企业、省级以上专精特新中小企业、省级战略性产业集群重点产业链“链主”企业等。"
        "（具体以每年申报通知为准）",
    ),
    # Thread: 设备金额 (行4-6，上下文关联)
    row(
        "省级企业技术中心项目单个项目新购置研发仪器设备总额的最低要求？",
        "单个项目新购置研发仪器设备(含配套软件，应为实施项目的新增设备，不包括生产设备、办公设备等)"
        "总额(不含税)不低于500万元。（具体以每年申报通知为准）",
    ),
    row(
        "不是600万元吗？",
        "根据当前的政策信息，省级企业技术中心项目的新购置研发仪器设备（含配套软件，不含税）总额的最低要求为500万元，并非600万元",
        ktype="结合上下文",
    ),
    row(
        "刚好500万元可以吗？",
        "是的，按照政策规定，省级企业技术中心项目的新购置研发仪器设备（含配套软件，不含税）总额的最低要求为500万元。"
        "因此，刚好500万元是符合最低要求的。",
        ktype="结合上下文",
    ),
    # Thread: 知识产权 (行7-8，上下文关联)
    row(
        "省级企业技术中心项目知识产权的最低要求？",
        "项目具有自主知识产权的创新成果，单个项目申请相关发明专利不少于3件。（具体以每年申报通知为准）",
    ),
    row(
        "只有2件可以申请吗",
        "根据政策规定，珠三角地区的省级企业技术中心项目申请相关发明专利等创新成果的最低要求是至少3件。"
        "因此，只有2件专利的情况下可能不符合申请条件。",
        ktype="结合上下文",
    ),
    row(
        "省级企业技术中心项目可申请奖励的范围？",
        "主要支持省级企业技术中心购置用于实施产业创新能力建设项目所需的新增设备（含配套软件），"
        "不包括生产设备、办公设备等，对应用于试验、检验检测、中试孵化、成果转化、产业应用（企业技术中心），"
        "平台建设、测试验证、中试孵化、成果转化（制造业创新中心）。（具体以每年申报通知为准）",
    ),
]

wb = Workbook()
ws = wb.active
ws.title = "问题表"

ws.append(HEADERS)
header_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
header_font = Font(bold=True, color="FFFFFF")
for col in range(1, len(HEADERS) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for row_data in ROWS:
    ws.append(row_data)

col_widths = [40, 60, 55, 55, 50, 10, 12, 12, 15]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

out = Path(__file__).parent / "questions.xlsx"
wb.save(out)
print(f"Saved {len(ROWS)} rows → {out}")

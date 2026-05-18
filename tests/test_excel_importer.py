from pathlib import Path

import pytest
from openpyxl import Workbook

from eval.importers.excel_importer import import_questions

FIXTURE = Path(__file__).parent / "fixtures" / "test_questions.xlsx"


def test_import_count():
    questions = import_questions(FIXTURE)
    assert len(questions) == 4


def test_ids_are_unique_and_sequential():
    questions = import_questions(FIXTURE)
    ids = [q["id"] for q in questions]
    assert ids == ["q_001", "q_002", "q_003", "q_004"]


def test_standalone_has_empty_history():
    questions = import_questions(FIXTURE)
    standalone = next(
        q
        for q in questions
        if q["knowledge_type"] == "文档" and "最低要求" in q["question"]
    )
    assert standalone["conversation_history"] == []


def test_first_followup_has_one_turn_of_history():
    questions = import_questions(FIXTURE)
    followup = next(q for q in questions if "600万元" in q["question"])
    assert len(followup["conversation_history"]) == 2
    assert followup["conversation_history"][0] == {
        "role": "user",
        "content": "最低要求是多少？",
    }
    assert followup["conversation_history"][1] == {
        "role": "assistant",
        "content": "500万元。",
    }


def test_second_followup_has_two_turns_of_history():
    questions = import_questions(FIXTURE)
    followup2 = next(q for q in questions if "只有2件" in q["question"])
    assert len(followup2["conversation_history"]) == 2
    assert followup2["conversation_history"][0]["content"] == "知识产权要求？"


def test_new_standalone_resets_thread():
    questions = import_questions(FIXTURE)
    second_standalone = next(q for q in questions if "知识产权" in q["question"])
    assert second_standalone["conversation_history"] == []


def test_multi_intent_mapping():
    questions = import_questions(FIXTURE)
    for q in questions:
        assert isinstance(q["is_multi_intent"], bool)


def test_prohibited_mapping():
    questions = import_questions(FIXTURE)
    for q in questions:
        assert isinstance(q["is_prohibited"], bool)
        assert q["is_prohibited"] is False


_HEADERS = [
    "用户输入的问题",
    "期望的回复",
    "关联的政策",
    "关联的文档",
    "申报入口",
    "是否多意图",
    "知识类型",
    "问题是否违禁",
    "备注",
]


def _make_xlsx(tmp_path, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(_HEADERS)
    for row in rows:
        ws.append(row)
    path = tmp_path / "fixture.xlsx"
    wb.save(path)
    return path


def test_empty_rows_do_not_skip_ids(tmp_path):
    path = _make_xlsx(
        tmp_path,
        [
            ["问题A", "答案A", "u1", "u2", "d.docx", "单意图", "文档", "正常", ""],
            [None] * 9,
            ["问题B", "答案B", "u1", "u2", "d.docx", "单意图", "文档", "正常", ""],
        ],
    )
    questions = import_questions(path)
    assert [q["id"] for q in questions] == ["q_001", "q_002"]


def test_empty_multi_intent_cell_defaults_to_false(tmp_path):
    """Empty multi-intent cell should mean '单意图' (safe default), not silently
    flag the row as multi-intent."""
    path = _make_xlsx(
        tmp_path,
        [["问题A", "答案A", "u1", "u2", "d.docx", "", "文档", "", ""]],
    )
    questions = import_questions(path)
    assert questions[0]["is_multi_intent"] is False
    assert questions[0]["is_prohibited"] is False


def test_unknown_multi_intent_raises_with_row_number(tmp_path):
    path = _make_xlsx(
        tmp_path,
        [["问题A", "答案A", "u1", "u2", "d.docx", "三意图", "文档", "正常", ""]],
    )
    with pytest.raises(ValueError, match=r"row 2.*is_multi_intent.*三意图"):
        import_questions(path)


def test_unknown_prohibited_raises_with_row_number(tmp_path):
    path = _make_xlsx(
        tmp_path,
        [
            ["问题A", "答案A", "u1", "u2", "d.docx", "单意图", "文档", "正常", ""],
            ["问题B", "答案B", "u1", "u2", "d.docx", "单意图", "文档", "未知状态", ""],
        ],
    )
    with pytest.raises(ValueError, match=r"row 3.*is_prohibited.*未知状态"):
        import_questions(path)


def test_unknown_knowledge_type_raises_with_row_number(tmp_path):
    path = _make_xlsx(
        tmp_path,
        [["问题A", "答案A", "u1", "u2", "d.docx", "单意图", "外部检索", "正常", ""]],
    )
    with pytest.raises(ValueError, match=r"row 2.*knowledge_type.*外部检索"):
        import_questions(path)


def test_named_sheet_selection(tmp_path):
    """import_questions should honor explicit sheet_name and not just fall back
    to the first/active sheet."""
    wb = Workbook()
    first = wb.active
    first.title = "junk"
    first.append(["irrelevant"])
    real = wb.create_sheet("data")
    real.append(_HEADERS)
    real.append(["问题A", "答案A", "u1", "u2", "d.docx", "单意图", "文档", "正常", ""])
    path = tmp_path / "multi.xlsx"
    wb.save(path)

    questions = import_questions(path, sheet_name="data")
    assert len(questions) == 1
    assert questions[0]["question"] == "问题A"
